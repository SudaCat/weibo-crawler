"""
Excel 写入模块
将爬虫结果写入 .xlsx 文件，支持批量写入和逐条增量追加
"""

from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

from loguru import logger

from config.settings import (
    EXCEL_HEADERS,
    EXCEL_FILENAME_PREFIX,
    EXCEL_COL_WIDTHS,
    RESULT_DIR,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# 样式常量（模块级复用）
# ============================================================
_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_FONT = Font(name="微软雅黑", size=10)
_DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_URL_COL_IDX = EXCEL_HEADERS.index("微博url") + 1  # 1-based


# ============================================================
# 创建带表头的 Workbook
# ============================================================
def create_result_workbook(
    output_path: Optional[Path] = None,
) -> Tuple[Workbook, Worksheet, Path]:
    """创建 Excel 工作簿并写入样式化的表头

    Args:
        output_path: 输出文件路径（可选，默认自动生成带时间戳的文件名）

    Returns:
        (workbook, worksheet, output_path)
    """
    if output_path is None:
        RESULT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{EXCEL_FILENAME_PREFIX}_{timestamp}.xlsx"
        output_path = RESULT_DIR / filename
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "爬虫结果"

    # 写表头
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # 设置列宽
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        width = EXCEL_COL_WIDTHS.get(header, 15)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    return wb, ws, output_path


# ============================================================
# 追加单行结果
# ============================================================
def append_result_row(ws: Worksheet, record: dict, row_idx: int) -> None:
    """向工作表中追加一行爬虫结果（带样式和超链接）

    Args:
        ws: 工作表对象
        record: 单条爬虫结果 dict
        row_idx: 行号（1-based，表头为第 1 行，数据从第 2 行开始）
    """
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        value = record.get(header, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = _DATA_FONT
        cell.alignment = _DATA_ALIGNMENT
        cell.border = _THIN_BORDER

        if col_idx == _URL_COL_IDX and value:
            cell.hyperlink = str(value)

    ws.row_dimensions[row_idx].height = 20


# ============================================================
# 批量写入
# ============================================================
def write_results(
    data: list[dict],
    output_path: Optional[Path] = None,
) -> Path:
    """将爬虫结果批量写入 Excel 文件

    Args:
        data: 爬虫结果列表
        output_path: 输出文件路径（可选）

    Returns:
        实际写入的文件路径
    """
    wb, ws, output_path = create_result_workbook(output_path)

    for i, record in enumerate(data, start=2):
        append_result_row(ws, record, i)

    ws.auto_filter.ref = ws.dimensions
    wb.save(str(output_path))
    logger.info(f"✅ 爬虫结果已写入: {output_path}（共 {len(data)} 条）")

    return output_path
